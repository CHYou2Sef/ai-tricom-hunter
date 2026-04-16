"""
╔══════════════════════════════════════════════════════════════════════════╗
║  agent.py  —  Core Orchestration Engine (Canonical Layout)               ║
║                                                                          ║
║  This file acts as the ORCHESTRATOR. It manages the pool of agents       ║
║  and routes logic to Specialized Agents (PhoneHunter, Enricher).         ║
╚══════════════════════════════════════════════════════════════════════════╝
"""

import os
import sys
import time
import random
import asyncio
from typing import List, Optional
from pathlib import Path

import config
from excel.reader import ExcelRow, read_excel
from excel.writer import save_results, save_subset_to_excel
from utils.logger import get_logger
from utils.metrics import PerformanceTracker
from utils.progress_tracker import ProgressTracker

# Canonical Agent Imports
from agents.phone_hunter import process_row
from agents.enricher import enrich_row

logger = get_logger(__name__)

# Initialize global progress tracker
progress = ProgressTracker(config.WORK_DIR / "active_processing.json")

# Counter for consecutive CAPTCHA blocks
_captcha_streak = 0
_captcha_lock = asyncio.Lock()

# Agent Pool for parallel workers
_agent_pool = asyncio.Queue()

async def init_agent_pool(count: int):
    """Pre-initialize a pool of browser hybrid engines."""
    logger.info(f"[AgentPool] Initializing {count} workers...")
    for i in range(count):
        from browser.hybrid_engine import HybridAutomationEngine
        agent = HybridAutomationEngine(worker_id=i+1)
        await agent.start_tier(config.HYBRID_DEFAULT_TIER)
        await _agent_pool.put(agent)

async def close_agent_pool():
    """Close all workers in the pool."""
    while not _agent_pool.empty():
        agent = await _agent_pool.get()
        await agent.close()

async def human_delay_async(min_sec: float = config.MIN_DELAY_SECONDS, max_sec: float = config.MAX_DELAY_SECONDS):
    delay = random.uniform(min_sec, max_sec)
    await asyncio.sleep(delay)

def sync_with_previous_results(rows: List[ExcelRow], filepath: str) -> int:
    """Synchronize with today's previous results to avoid redundant scraping."""
    from pathlib import Path
    import openpyxl
    import datetime
    from utils.column_detector import detect_columns

    orig_path = Path(filepath)
    input_folder = orig_path.parent.name
    out_dir = config.get_output_dir(input_folder)
    date_str = datetime.date.today().strftime("%Y-%m-%d")
    fusion_path = out_dir / f"{input_folder}_{date_str}.xlsx"
    
    file_key = f"{input_folder}/{orig_path.name}"
    completed_rows = progress.get_completed_rows(file_key)
    if completed_rows:
        for r in rows:
            if r.row_index in completed_rows:
                r.status = "DONE"
    
    if not fusion_path.exists(): return 0
    sync_count = 0
    try:
        wb = openpyxl.load_workbook(fusion_path, data_only=True)
        ws = wb.active
        headers = [str(cell.value) for cell in ws[1] if cell.value is not None]
        col_map = {name: i for i, name in enumerate(headers)}
        out_mapping = detect_columns(headers)
        
        lookup = {}
        def get_val(row_data, concept):
            col_name = out_mapping.get(concept)
            if col_name and col_name in col_map:
                val = row_data[col_map[col_name]].value
                return str(val).strip() if val is not None else None
            return None

        for row_data in ws.iter_rows(min_row=2):
            if all(c.value is None for c in row_data): continue
            dummy = ExcelRow(raw={}, row_index=0, mapping={})
            dummy.nom = get_val(row_data, "raison_sociale")
            dummy.adresse = get_val(row_data, "adresse")
            dummy.siren = get_val(row_data, "siren") or get_val(row_data, "siret")
            
            fp = dummy.get_fingerprint()
            lookup[fp] = {
                "status": (get_val(row_data, "Etat") or "DONE").upper(),
                "phone": get_val(row_data, "AI_Phone") or get_val(row_data, "telephone"),
                "agent_phone": get_val(row_data, "AI_Agent_Phone")
            }
        
        for r in rows:
            fp = r.get_fingerprint()
            if fp in lookup:
                entry = lookup[fp]
                if entry["status"] in ("DONE", "NO TEL", "SKIP"):
                    r.status, r.phone, r.agent_phone = entry["status"], entry["phone"], entry["agent_phone"]
                    sync_count += 1
        wb.close()
    except Exception as e:
        logger.warning(f"⚠️ [Sync] Link failed: {e}")
    return sync_count

async def _worker_process_row(row: ExcelRow, sem: asyncio.Semaphore, save_lock: asyncio.Lock, rows: List[ExcelRow], filepath: str, tracker: PerformanceTracker, idx: int, total: int) -> ExcelRow:
    async with sem:
        agent = await _agent_pool.get()
        worker_id = getattr(agent, "worker_id", "?")
        row_start = time.perf_counter()

        try:
            # 1. PEFORM SEARCH (Specialized Agent)
            await process_row(row, agent, idx=idx, total=total)
            
            # 2. PERFORM ENRICHMENT (Specialized Agent)
            if row.status != "SKIP":
                await asyncio.to_thread(enrich_row, row)

            # Mark progress
            input_folder = Path(filepath).parent.name
            file_key = f"{input_folder}/{Path(filepath).name}"
            progress.mark_row_done(file_key, row.row_index)

        except Exception as e:
            logger.error(f"[Worker-{worker_id}] Error row #{row.row_index}: {e}")
        finally:
            await _agent_pool.put(agent)

        elapsed = time.perf_counter() - row_start
        tracker.track_row(elapsed, row.status)

        if idx % config.SAVE_INTERVAL == 0 or idx == total:
            async with save_lock:
                await asyncio.to_thread(save_results, rows, filepath)

    return row

async def process_file_async(filepath: str) -> None:
    logger.info(f"[Agent] Starting file: {os.path.basename(filepath)}")
    rows, _ = await asyncio.to_thread(read_excel, filepath)
    if not rows: return

    tracker = PerformanceTracker()
    tracker.start_file_processing()

    await asyncio.to_thread(sync_with_previous_results, rows, filepath)
    rows_to_process = [r for r in rows if r.status != "DONE"] if config.REPROCESS_FAILED_ROWS else [r for r in rows if r.status not in ("DONE", "NO TEL", "SKIP")]

    total = len(rows_to_process)
    if total > 0:
        save_lock, sem = asyncio.Lock(), asyncio.Semaphore(config.MAX_CONCURRENT_WORKERS)
        tasks = [asyncio.create_task(_worker_process_row(r, sem, save_lock, rows, filepath, tracker, i, total)) for i, r in enumerate(rows_to_process, 1)]
        await asyncio.gather(*tasks)
    
    await asyncio.to_thread(save_results, rows, filepath)
    tracker.end_file_processing()
    await finalize_file_processing(rows, filepath)

async def finalize_file_processing(rows: List[ExcelRow], original_filepath: str) -> None:
    orig_path = Path(original_filepath)
    success_rows = [r for r in rows if r.phone or r.status == "DONE"]
    retry_rows = [r for r in rows if r.status in ("NO TEL", "SKIP", "PENDING")]

    if success_rows:
        config.OUTPUT_SUCCEED_DIR.mkdir(parents=True, exist_ok=True)
        await asyncio.to_thread(save_subset_to_excel, success_rows, config.OUTPUT_SUCCEED_DIR / orig_path.name)
    if retry_rows:
        config.OUTPUT_FAILED_DIR.mkdir(parents=True, exist_ok=True)
        await asyncio.to_thread(save_subset_to_excel, retry_rows, config.OUTPUT_FAILED_DIR / orig_path.name)

    progress.clear_file(f"{orig_path.parent.name}/{orig_path.name}")
    try: os.remove(original_filepath)
    except: pass
