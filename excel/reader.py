"""
╔══════════════════════════════════════════════════════════════════════════╗
║  excel/reader.py                                                         ║
║                                                                          ║
║  Reads Excel files (.xlsx/.xls) and extracts rows into clean dicts.     ║
║  Uses column_detector.py to map column names automatically.             ║
║                                                                          ║
║  BEGINNER NOTE:                                                          ║
║    We use the `openpyxl` library to read Excel files.                   ║
║    Each row in the Excel becomes a Python dictionary:                   ║
║      { "raison_sociale": "...", "adresse": "...", "siren": "..." }      ║
╚══════════════════════════════════════════════════════════════════════════╝
"""

import csv
import io
import os
import json
import re
import asyncio
from typing import List, Optional, Tuple

import openpyxl

import config
from utils.column_detector import detect_columns, validate_mapping
from utils.llm_parser import detect_columns_with_llm
from utils.logger import get_logger

logger = get_logger(__name__)


class ExcelRow:
    """
    Represents a single cleaned row from the Excel file.

    Attributes:
        raw          : The original row data as a dict {column_name: value}
        row_index    : The row number in the Excel file (1-based, header = 1)
        nom          : Raison Sociale (company name) — may be None
        adresse      : Full address — may be None
        siren        : SIREN number — may be None
        siret        : SIRET number — may be None
        search_type  : "RS_ADR" if nom+adresse available, "SIREN_ADR" otherwise
        status       : Processing status ("PENDING", "DONE", "NO TEL", "ERROR")
        phone        : Extracted phone number (None until found)
        agent_phone  : Agent/commercial phone (None until found)
    """

    def __init__(self, raw: dict, row_index: int, mapping: dict):
        self.raw       = raw
        self.row_index = row_index

        # ── Extract key fields using the column mapping ──
        def get(concept: str) -> Optional[str]:
            col = mapping.get(concept)
            if col and col in raw:
                val = raw[col]
                s = str(val).strip() if val is not None else ""
                return s if s and s.lower() not in ("none", "nan", "") else None
            return None

        self.nom     = get("raison_sociale")
        self.adresse = get("adresse")
        self.siren   = get("siren") or get("siret")

        if self.nom and self.adresse:
            self.search_type = "RS_ADR"
        elif self.siren and self.adresse:
            self.search_type = "SIREN_ADR"
        elif self.nom:
             # If we only have name, we can still try to search
             self.search_type = "RS_ADR"
        elif self.siren:
             self.search_type = "SIREN_ADR"
        else:
            self.search_type = "SKIP"

        self.status = ""
        etat_val = get("Etat") or get("stat")
        if etat_val and etat_val.upper() == "DONE":
            self.status = "DONE"
        elif etat_val and "NO" in etat_val.upper() and ("TEL" in etat_val.upper() or "PHONE" in etat_val.upper()):
            self.status = "NO TEL"
        
        self.phone = get("telephone") or get("phone") or get("téléphone") or get("__phone")
        self.agent_phone = get("agent_phone") or get("__agent_phone")

        self.enriched_fields: dict = {}
        self.raw_ai_responses: list = []
        self.search_queries_used: list = []
        self.processing_start_ts: float = 0.0
        self.processing_end_ts: float = 0.0
        self.captcha_hits: int = 0

    def get_fingerprint(self) -> str:
        """Create a unique string key for this row to track progress."""
        if self.siren and len(self.siren) >= 9:
            return f"SIREN:{self.siren}"
        
        n = str(self.nom or "").strip().lower()
        a = str(self.adresse or "").strip().lower()
        n = re.sub(r'[^a-z0-9]', '', n)
        a = re.sub(r'[^a-z0-9]', '', a)
        return f"NA:{n}|{a}"

    def get_search_name(self) -> str:
        """Return the best identifier for searching (nom or SIREN)."""
        return self.nom if self.nom else (self.siren or "")

    def to_dict(self) -> dict:
        """
        Convert this row to a flat dictionary for JSON/Excel output.
        Includes all original columns PLUS our new fields.
        """
        result = dict(self.raw)   # Start with original data
        result.update({
            "__row_index":    self.row_index,
            "__search_type":  self.search_type,
            "__phone":        self.phone or "",
            "__agent_phone":  self.agent_phone or "",
            "__status":       self.status,
        })
        return result

    def __repr__(self) -> str:
        return (
            f"<ExcelRow #{self.row_index} | "
            f"type={self.search_type} | "
            f"nom='{self.nom}' | "
            f"status={self.status}>"
        )


def find_header_row(ws, max_rows=15) -> int:
    """
    Scanne les premières lignes du fichier Excel pour trouver l'index 
    de la ligne contenant les titres, basé sur un score de mots-clés.
    """
    target_keywords = {
        "siren", "nom", "dénomination", "adresse",
        "forme juridique", "activité", "immatriculation", "statut"
    }
    
    best_row_idx = 1
    max_matches = 0

    for row_idx, row_cells in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
        row_values = [str(val).lower() for val in row_cells if val is not None]
        matches = sum(1 for word in target_keywords if any(word in val for val in row_values))
        
        if matches > max_matches:
            max_matches = matches
            best_row_idx = row_idx

    return best_row_idx


def read_excel(filepath: str) -> Tuple[List[ExcelRow], dict]:
    """
    Read an Excel file and return:
        1. A list of ExcelRow objects (one per data row)
        2. The column mapping dict detected by column_detector

    Args:
        filepath : Full path to the .xlsx or .xls file

    Returns:
        (rows, mapping)  where rows is List[ExcelRow]

    Raises:
        FileNotFoundError : if the file doesn't exist
        ValueError        : if no recognizable columns found
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    logger.info(f"[Reader] Opening: {os.path.basename(filepath)}")

    # ── Route CSV files separately ──
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".csv":
        return _read_csv(filepath)
    if ext == ".json":
        return _read_json(filepath)

    # data_only=True  → read calculated cell values, not formulas
    # read_only=False → we need full access
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active   # Use the first (active) sheet

    # ── Auto-detect header row ──
    header_idx = find_header_row(ws)
    logger.info(f"[Reader] Auto-detected headers at row {header_idx}")

    # ── Extract headers from the detected row ──
    headers = []
    for cell in ws[header_idx]:
        value = cell.value
        # Nettoyage des noms de colonnes (suppression des sauts de ligne, espaces en trop)
        cleaned_val = str(value).replace('\n', ' ').strip() if value is not None else ""
        headers.append(cleaned_val)

    logger.info(f"[Reader] Found {len(headers)} columns: {headers}")

    # ── Auto-detect column mapping ──
    mapping = detect_columns(headers)
    validation = validate_mapping(mapping)

    logger.info(f"[Reader] Column heuristic mapping: { {k:v for k,v in mapping.items() if v} }")

    if not validation["can_search_rs"] and not validation["can_search_siren"]:
        logger.warning("[Reader] Heuristics failed to detect sufficient columns.")
        # ── LLM Fallback Mapping ──
        sample_rows_raw = list(ws.iter_rows(min_row=header_idx + 1, max_row=header_idx + 3, values_only=True))
        if sample_rows_raw:
            llm_mapping = asyncio.run(detect_columns_with_llm(headers, list(sample_rows_raw)))
            if llm_mapping:
                mapping.update({k: v for k, v in llm_mapping.items() if v})
                validation = validate_mapping(mapping)
                logger.info(f"[Reader] LLM Fallback applied: { {k:v for k,v in mapping.items() if v} }")

    if not validation["can_search_rs"] and not validation["can_search_siren"]:
        logger.warning(
            "[Reader] WARNING: Could not detect enough columns to search even with LLM. "
            "Rows will be marked SKIP. Check column names."
        )

    # ── Read data rows (starting right after the header) ──
    rows: List[ExcelRow] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_idx + 1, values_only=True), start=header_idx + 1):
        # Skip completely empty rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        # Build a dict { column_header: cell_value } for this row
        raw = {}
        for col_idx, header in enumerate(headers):
            if not header: continue # Skip completely empty ghost columns
            if col_idx < len(row):
                raw[header] = row[col_idx]
            else:
                raw[header] = None

        # ── Filtrage des Radiées (if statutory column exists) ──
        statut_keys = [k for k in raw.keys() if "statut" in k.lower() or "état" in k.lower() or "etat" in k.lower()]
        is_radiee = False
        for sk in statut_keys:
            val = str(raw[sk]).lower()
            if "radiée" in val or "radiee" in val:
                is_radiee = True
                break
        
        if is_radiee:
            logger.debug(f"[Reader] Row {row_idx} skipped (Statut: Radiée)")
            continue

        excel_row = ExcelRow(raw=raw, row_index=row_idx, mapping=mapping)
        
        # ── Normalisation SIREN (pad with zeros, remove .0, remove spaces) ──
        if excel_row.siren:
            excel_row.siren = str(excel_row.siren).replace(".0", "").replace(" ", "").zfill(9)

        rows.append(excel_row)

    logger.info(
        f"[Reader] Loaded {len(rows)} data rows "
        f"({sum(1 for r in rows if r.search_type=='RS_ADR')} RS_ADR, "
        f"{sum(1 for r in rows if r.search_type=='SIREN_ADR')} SIREN_ADR, "
        f"{sum(1 for r in rows if r.search_type=='SKIP')} SKIP)"
    )

    wb.close()
    return rows, mapping


def _read_csv(filepath: str) -> Tuple[List[ExcelRow], dict]:
    """
    Read a .csv file and return the same (rows, mapping) as read_excel.
    Auto-detects delimiter: comma, semicolon, tab, or pipe (French exports use ';').
    Supports fallback encodings (utf-8, latin-1, mac_roman, utf-16).
    """
    logger.info(f"[Reader] CSV mode: {os.path.basename(filepath)}")

    # 1. Determine safe encoding
    encodings_to_try = ["utf-8-sig", "latin-1", "mac_roman", "utf-16"]
    best_encoding = "utf-8-sig"
    sample = ""
    for enc in encodings_to_try:
        try:
            with open(filepath, "r", encoding=enc) as f:
                sample = f.read(4096)
                best_encoding = enc
                break
        except UnicodeDecodeError:
            continue

    if not sample:
        logger.warning(f"[Reader] All encodings failed for {filepath}. Falling back to utf-8 with replacement.")
        best_encoding = "utf-8-sig"
        with open(filepath, "r", encoding=best_encoding, errors="replace") as f:
            sample = f.read(4096)

    # 2. Sniff delimiter
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";"  # French CSV default (e.g. data.gouv.fr exports)
    logger.info(f"[Reader] CSV delimiter: '{delimiter}' | Encoding: '{best_encoding}'")

    # 3. Read data
    rows_raw: List[dict] = []
    with open(filepath, "r", encoding=best_encoding, errors="replace") as f:
        # --- 3. Header Detection (Bug Fix for Headless CSVs) ---
        sample_reader = csv.reader(io.StringIO(sample), delimiter=delimiter)
        first_row = next(sample_reader, [])
        
        # Heuristic: if first row contains SIREN pattern or many digits, it's DATA
        is_data = any(re.fullmatch(r'\d{9}|\d{3}\s\d{3}\s\d{3}', str(v).strip()) for v in first_row)
        
        # Count keywords
        keywords = ["nom", "siren", "siret", "adresse", "date", "status", "etat", "denomination", "rs"]
        header_score = sum(1 for v in first_row if any(kw in str(v).lower() for kw in keywords))
        
        if is_data or header_score == 0:
            # Headless file! Use positional headers and DON'T skip first row
            f.seek(0)
            raw_headers = [f"col_{i+1}" for i in range(len(first_row))]
            reader = csv.DictReader(f, fieldnames=raw_headers, delimiter=delimiter)
            # We don't need to skip any row here because DictReader(fieldnames=...) treats first row as data
        else:
            # Normal file with headers
            f.seek(0)
            reader = csv.DictReader(f, delimiter=delimiter)
            raw_headers = list(reader.fieldnames or [])
        for row in reader:
            if all(not str(v).strip() for v in row.values()):
                continue
            rows_raw.append(dict(row))

    if not rows_raw:
        logger.warning("[Reader] CSV file appears empty.")
        return [], {}

    # Clean header names (protect against None)
    headers = [h.replace("\n", " ").strip() for h in raw_headers if h]
    logger.info(f"[Reader] CSV columns ({len(headers)}): {headers}")

    mapping    = detect_columns(headers)
    validation = validate_mapping(mapping)
    logger.info(f"[Reader] CSV column heuristic mapping: { {k:v for k,v in mapping.items() if v} }")

    if not validation["can_search_rs"] and not validation["can_search_siren"]:
        logger.warning("[Reader] CSV Heuristics failed to detect sufficient columns.")
        sample_rows_raw = [list(r.values()) for r in rows_raw[:3]]
        if sample_rows_raw:
            llm_mapping = asyncio.run(detect_columns_with_llm(headers, sample_rows_raw))
            if llm_mapping:
                mapping.update({k: v for k, v in llm_mapping.items() if v})
                validation = validate_mapping(mapping)
                logger.info(f"[Reader] CSV LLM Fallback applied: { {k:v for k,v in mapping.items() if v} }")

    if not validation["can_search_rs"] and not validation["can_search_siren"]:
        logger.warning(
            "[Reader] WARNING: Could not detect enough columns. "
            "Rows will be marked SKIP. Check column names."
        )

    rows: List[ExcelRow] = []
    for row_idx, raw in enumerate(rows_raw, start=2):  # row 1 = header
        # Filter Radiées
        statut_keys = [k for k in raw if k and ("statut" in k.lower() or "état" in k.lower() or "etat" in k.lower())]
        if any("radi" in str(raw[sk]).lower() for sk in statut_keys):
            logger.debug(f"[Reader] CSV row {row_idx} skipped (Radiée)")
            continue

        excel_row = ExcelRow(raw=raw, row_index=row_idx, mapping=mapping)

        if excel_row.siren:
            excel_row.siren = str(excel_row.siren).replace(".0", "").replace(" ", "").zfill(9)

        rows.append(excel_row)

    logger.info(
        f"[Reader] CSV loaded {len(rows)} rows "
        f"({sum(1 for r in rows if r.search_type=='RS_ADR')} RS_ADR, "
        f"{sum(1 for r in rows if r.search_type=='SIREN_ADR')} SIREN_ADR, "
        f"{sum(1 for r in rows if r.search_type=='SKIP')} SKIP)"
    )
    return rows, mapping


def _read_json(filepath: str) -> Tuple[List[ExcelRow], dict]:
    """
    Read a .json file containing a list of dictionaries.
    Compatible with industrial chunks and bulk exports.
    """
    logger.info(f"[Reader] JSON mode: {os.path.basename(filepath)}")
    
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        logger.error(f"[Reader] Failed to parse JSON: {e}")
        return [], {}

    if not isinstance(data, list):
        # Support single object case
        if isinstance(data, dict):
            data = [data]
        else:
            logger.error("[Reader] JSON content is not a list of objects.")
            return [], {}

    if not data:
        return [], {}

    # Extract all unique keys from all objects to build headers
    headers_set = set()
    for item in data:
        if isinstance(item, dict):
            headers_set.update(item.keys())
    
    headers = sorted(list(headers_set))
    mapping = detect_columns(headers)
    validation = validate_mapping(mapping)
    
    rows: List[ExcelRow] = []
    for idx, raw in enumerate(data, start=2):
        if not isinstance(raw, dict):
            continue
        excel_row = ExcelRow(raw=raw, row_index=idx, mapping=mapping)
        if excel_row.siren:
            excel_row.siren = excel_row.siren.replace(".0", "").zfill(9)
        rows.append(excel_row)

    logger.info(
        f"[Reader] JSON loaded {len(rows)} rows "
        f"({sum(1 for r in rows if r.search_type=='RS_ADR')} RS_ADR, "
        f"{sum(1 for r in rows if r.search_type=='SIREN_ADR')} SIREN_ADR, "
        f"{sum(1 for r in rows if r.search_type=='SKIP')} SKIP)"
    )
    return rows, mapping
