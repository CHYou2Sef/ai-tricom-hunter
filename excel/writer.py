"""
excel/writer.py

Saves processed rows to EXCEL and rich JSON.
Excel: original columns + AI_Phone + AI_Agent_Phone + Etat
JSON:  full audit record per row.
"""

import os
import json
from pathlib import Path
from openpyxl import Workbook
import config
from utils.logger import get_logger

logger = get_logger(__name__)

# save_rich_json entirely removed to stop .json creation

def save_results(rows: list, original_filepath: str) -> None:
    """
    Save results to EXCEL.
    Excel: original columns + AI_Phone + AI_Agent_Phone + Etat
           + one column per enriched field that was filled.
    
    Fusion: if a file with the same date exists in Extraction_{folder}, append to it.
    """
    import datetime
    import openpyxl
    
    orig_path = Path(original_filepath)
    filename    = orig_path.name
    name_no_ext = orig_path.stem
    
    # Get the name of the folder where the input file was found
    input_folder = orig_path.parent.name
    out_dir = config.get_output_dir(input_folder)
    
    date_str = datetime.date.today().strftime("%Y-%m-%d")
    fusion_path = out_dir / f"{input_folder}_{date_str}.xlsx"
    
    # ── Discover Enriched Headers ──
    enriched_keys = set()
    for r in rows:
        enriched_keys.update(getattr(r, "enriched_fields", {}).keys())
    enriched_keys = sorted(enriched_keys)
    
    base_headers = list(rows[0].raw.keys()) if rows else []
    ai_status_header = config.STATUS_COLUMN_NAME
    
    # AI Results columns (only add if not in base headers to avoid duplicates)
    candidate_ai_headers = ["AI_Phone", "AI_Agent_Phone", ai_status_header]
    ai_headers = [h for h in candidate_ai_headers if h not in base_headers]
    
    # Enriched Fields columns
    candidate_enr_headers = [f"AI_{k.upper()}" for k in enriched_keys]
    enr_headers = [h for h in candidate_enr_headers if h not in base_headers]
    
    all_headers = base_headers + ai_headers + enr_headers

    # ── Excel Workbook ──
    existing_fps = set()
    try:
        if fusion_path.exists():
            wb = openpyxl.load_workbook(fusion_path)
            ws = wb.active
            # Read existing headers to map correctly and not shift columns
            existing_headers = [str(cell.value) for cell in ws[1] if cell.value is not None]
            
            # --- Progress Optimization: Pre-scan for duplicates ---
            from utils.column_detector import detect_columns
            out_mapping = detect_columns(existing_headers)
            col_map = {name: i for i, name in enumerate(existing_headers)}
            
            def get_cell_val(row_cells, concept):
                col_name = out_mapping.get(concept)
                if col_name and col_name in col_map:
                    val = row_cells[col_map[col_name]].value
                    return str(val).strip() if val is not None else None
                return None

            for row_cells in ws.iter_rows(min_row=2):
                o_nom = get_cell_val(row_cells, "raison_sociale")
                o_adr = get_cell_val(row_cells, "adresse")
                o_siren = get_cell_val(row_cells, "siren") or get_cell_val(row_cells, "siret")
                
                # Fingerprint reconstruction
                import re
                n = str(o_nom or "").strip().lower()
                a = str(o_adr or "").strip().lower()
                n = re.sub(r'[^a-z0-9]', '', n)
                a = re.sub(r'[^a-z0-9]', '', a)
                if o_siren:
                    o_siren = str(o_siren).replace(".0", "").replace(" ", "").zfill(9)
                if o_siren and len(o_siren) >= 9:
                    fp = f"SIREN:{o_siren}"
                else:
                    fp = f"NA:{n}|{a}"
                existing_fps.add(fp)

            # Expanded headers logic
            new_cols_found = [h for h in all_headers if h not in existing_headers]
            if new_cols_found:
                next_col = len(existing_headers) + 1
                for new_h in new_cols_found:
                    ws.cell(row=1, column=next_col).value = new_h
                    existing_headers.append(new_h)
                    next_col += 1
                logger.info(f"[Writer] Expanded headers with: {new_cols_found}")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            ws.append(all_headers)
            existing_headers = all_headers
    except Exception as exc:
        logger.error(f"[Writer] CRITICAL: File {fusion_path.name} is CORRUPTED. Skipping. Error: {exc}")
        # Automatically move corrupted file to a safe subfolder
        recovery_dir = fusion_path.parent / "CORRUPTED_FILES"
        recovery_dir.mkdir(parents=True, exist_ok=True)
        import os
        try:
            os.rename(fusion_path, recovery_dir / fusion_path.name)
        except: pass
        return
        
    rows_added = 0
    for r in rows:
        # Check if row is already in the daily fusion file
        if r.get_fingerprint() in existing_fps:
            continue

        # Determine Status accurately
        status_to_write = ""
        if r.phone:
            status_to_write = "Done"
        elif r.status == "NO TEL":
            status_to_write = "No Tel"
        elif r.status == "DONE":
            status_to_write = "Done"

        # Map data by column name to prevent sliding
        row_dict = {}
        # 1. Fill base data
        for h in base_headers:
            row_dict[h] = r.raw.get(h, "")
        # 2. Fill AI Core
        row_dict["AI_Phone"] = r.phone or ""
        row_dict["AI_Agent_Phone"] = r.agent_phone or ""
        row_dict[ai_status_header] = status_to_write
        # 3. Fill Enriched
        for k in enriched_keys:
            info = getattr(r, "enriched_fields", {}).get(k, {})
            row_dict[f"AI_{k.upper()}"] = info.get("value", "")

        # Write ordered data based on actual worksheet headers
        ordered_data = []
        for h in existing_headers:
            val = row_dict.get(h, "")
            if isinstance(val, (list, dict)):
                val = str(val)
            ordered_data.append(val)
            
        ws.append(ordered_data)
        rows_added += 1

    if rows_added > 0:
        wb.save(fusion_path)
        logger.info(f"[Writer] Results fused/saved: {fusion_path} (+{rows_added} new rows)")
    else:
        logger.info(f"[Writer] All rows already exist in the fusion file. No changes saved.")


def save_subset_to_excel(rows: list, target_path: Path) -> None:
    """
    Save a list of ExcelRow objects to a NEW Excel file.
    Uses dictionary mapping to avoid column sliding.
    """
    import openpyxl
    if not rows:
        return

    # Discover Enriched Headers
    enriched_keys = set()
    for r in rows:
        enriched_keys.update(getattr(r, "enriched_fields", {}).keys())
    enriched_keys = sorted(enriched_keys)

    base_headers = list(rows[0].raw.keys())
    ai_status_header = config.STATUS_COLUMN_NAME
    
    # AI Results columns
    candidate_ai_headers = ["AI_Phone", "AI_Agent_Phone", ai_status_header]
    ai_headers = [h for h in candidate_ai_headers if h not in base_headers]
    
    # Enriched Fields columns
    candidate_enr_headers = [f"AI_{k.upper()}" for k in enriched_keys]
    enr_headers = [h for h in candidate_enr_headers if h not in base_headers]
    
    all_headers = base_headers + ai_headers + enr_headers

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(all_headers)

    for r in rows:
        status_to_write = ""
        if r.phone:
            status_to_write = "Done"
        elif r.status == "NO TEL":
            status_to_write = "No Tel"
        elif r.status == "DONE":
            status_to_write = "Done"

        # Map data by column name
        row_dict = {}
        for h in base_headers:
            row_dict[h] = r.raw.get(h, "")
        
        row_dict["AI_Phone"] = r.phone or ""
        row_dict["AI_Agent_Phone"] = r.agent_phone or ""
        row_dict[ai_status_header] = status_to_write
        
        for k in enriched_keys:
            info = getattr(r, "enriched_fields", {}).get(k, {})
            row_dict[f"AI_{k.upper()}"] = info.get("value", "")
            
        ordered_data = []
        for h in all_headers:
            val = row_dict.get(h, "")
            if isinstance(val, (list, dict)):
                val = str(val)
            ordered_data.append(val)
            
        ws.append(ordered_data)

    target_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target_path)
    logger.info(f"[Writer] Subset saved: {target_path}")
